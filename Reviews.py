import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import signal

def save_reviews_and_exit(driver, workbook, file_path, total_reviews_collected):
    workbook.save(filename=file_path)
    print(f"\nTotal number of collected reviews: {total_reviews_collected}")
    print(f"The Excel file has been saved at: {os.path.abspath(file_path)}")
    driver.quit()
    exit(0)

def signal_handler(sig, frame, driver, workbook, file_path, total_reviews_collected):
    print('\nYou pressed Ctrl+C!')
    save_reviews_and_exit(driver, workbook, file_path, total_reviews_collected)

def get_user_input():
    user_input_item = input("Enter the item you want to scrape: ")
    max_pages = input("Enter the number of pages to collect reviews up to (leave blank to scrape all pages): ")
    max_pages = int(max_pages) if max_pages else float('inf')
    return user_input_item, max_pages

def initialize_driver():
    os.environ["PATH"] += os.pathsep + 'C:\\chromedriver.exe'
    chrome_options = Options()
    chrome_options.add_argument("--incognito")
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def get_base_url(user_input_item):
    base_url = f"https://www.daraz.com.bd/catalog/?q={user_input_item}&_keyori=ss&from=input&spm=a2a0e.home.search.go.285012f7rLpOTH"
    return base_url

def initialize_workbook(file_path):
    try:
        workbook = load_workbook(filename=file_path)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Reviews'
        worksheet.append(['Item', 'Product Name', 'Brand Name', 'Total Ratings', 'Price After Discount', 'Actual Price',
                          'Discount Percentage', 'Review', '5 Star %', '4 Star %', '3 Star %', '2 Star %', '1 Star %',
                          'Chat Response Rate', 'Ship on Time', 'Sold By', 'Cash on Delivery', 'Warranty'])
    if worksheet.max_row == 1 and all(cell.value is None for cell in worksheet[1]):
        worksheet.append(['Item', 'Product Name', 'Brand Name', 'Total Ratings', 'Price After Discount', 'Actual Price',
                          'Discount Percentage', 'Review', '5 Star %', '4 Star %', '3 Star %', '2 Star %', '1 Star %',
                          'Chat Response Rate', 'Ship on Time', 'Sold By', 'Cash on Delivery', 'Warranty'])
    return workbook, worksheet

def scroll_to_reviews(driver):
    current_height = 0
    while True:
        driver.execute_script(f"window.scrollTo(0, {current_height});")
        time.sleep(1)
        reviews = driver.find_elements(By.CLASS_NAME, "review-content-sl")
        if reviews:
            return reviews
        current_height += 800
        if current_height >= driver.execute_script("return document.body.scrollHeight;") // 2:
            return None

def collect_product_info(driver):
    product_name = ""
    brand_name = ""
    total_ratings = 0
    price_after_discount = 0
    actual_price = 0
    discount_percentage = 0
    chat_response_rate = 0
    ship_on_time = "Not enough data"
    sold_by = ""
    cash_on_delivery = "No"
    warranty = "Warranty not available"
    try:
        product_name_element = driver.find_element(By.CSS_SELECTOR, ".pdp-mod-product-badge-title")
        product_name = product_name_element.text
    except:
        pass
    try:
        brand_name_element = driver.find_element(By.CSS_SELECTOR, ".pdp-product-brand__brand-link")
        brand_name = brand_name_element.text
    except:
        pass
    try:
        total_ratings_element = driver.find_element(By.CSS_SELECTOR, ".pdp-review-summary__link")
        total_ratings = int(total_ratings_element.text.split()[0])
    except:
        pass
    try:
        price_after_discount_element = driver.find_element(By.CSS_SELECTOR, ".pdp-price_type_normal")
        price_after_discount = float(price_after_discount_element.text.replace("৳", "").replace(",", ""))
    except:
        pass
    try:
        actual_price_element = driver.find_element(By.CSS_SELECTOR, ".pdp-price_type_deleted")
        actual_price = float(actual_price_element.text.replace("৳", "").replace(",", ""))
    except:
        pass
    try:
        discount_percentage_element = driver.find_element(By.CSS_SELECTOR, ".pdp-product-price__discount")
        discount_percentage = abs(int(discount_percentage_element.text.replace("-", "").replace("%", "")))
    except:
        pass
    try:
        chat_response_rate_element = driver.find_element(By.CSS_SELECTOR, ".seller-info-value")
        chat_response_rate = int(chat_response_rate_element.text.replace("%", ""))
    except:
        pass
    try:
        ship_on_time_element = driver.find_element(By.XPATH, "//div[contains(@class, 'info-content')]/div[contains(text(), 'Ship on Time')]/following-sibling::div")
        ship_on_time = ship_on_time_element.text
    except:
        pass
    try:
        sold_by_element = driver.find_element(By.CSS_SELECTOR, ".seller-name__detail-name")
        sold_by = sold_by_element.text
    except:
        pass
    try:
        cash_on_delivery_element = driver.find_element(By.XPATH, "//div[contains(@class, 'delivery-option-item_type_COD')]")
        cash_on_delivery = "Yes"
    except:
        pass
    try:
        warranty_element = driver.find_element(By.XPATH, "//div[contains(@class, 'delivery-option-item_type_noWarranty')]")
        warranty = "Warranty not available"
    except:
        warranty = "Warranty available"
    return product_name, brand_name, total_ratings, price_after_discount, actual_price, discount_percentage, chat_response_rate, ship_on_time, sold_by, cash_on_delivery, warranty

def collect_star_percentages(driver):
    star_percentages = []
    try:
        star_percentage_elements = driver.find_elements(By.CSS_SELECTOR, ".review-info-right .percent")
        for element in star_percentage_elements:
            percentage = int(element.text.strip('%'))
            star_percentages.append(percentage)
    except:
        pass
    return star_percentages

def collect_reviews(driver, user_input_item, worksheet, max_pages, total_reviews_collected):
    current_page = 1
    column_headers = [cell.value for cell in worksheet[1]]
    while current_page <= max_pages:
        try:
            reviews = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "review-content-sl"))
            )
        except:
            break
        product_name, brand_name, total_ratings, price_after_discount, actual_price, discount_percentage, chat_response_rate, ship_on_time, sold_by, cash_on_delivery, warranty = collect_product_info(driver)
        star_percentages = collect_star_percentages(driver)
        for review in reviews:
            review_text = review.text
            row_data = [user_input_item, product_name, brand_name, total_ratings, price_after_discount, actual_price,
                        discount_percentage, review_text] + star_percentages + [chat_response_rate, ship_on_time, sold_by, cash_on_delivery, warranty]
            worksheet.append([value for value, header in zip(row_data, column_headers)])
            total_reviews_collected += 1
        print(f"\rCollecting reviews on page {current_page} of the item's review. {len(reviews)} reviews collected. Total number of reviews collected: {total_reviews_collected}", end="", flush=True)
        if current_page == max_pages:
            break
        try:
            pagination_items = driver.find_elements(By.CSS_SELECTOR, ".ant-pagination-item")
            next_page_link = None
            for item in pagination_items:
                if item.get_attribute("title") == str(current_page + 1):
                    next_page_link = item.find_element(By.TAG_NAME, "a")
                    break
            if next_page_link:
                next_page_link.click()
                current_page += 1
                time.sleep(2)
            else:
                break
        except:
            break
    return total_reviews_collected

def main():
    user_input_item, max_pages = get_user_input()
    driver = initialize_driver()
    base_url = get_base_url(user_input_item)
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_name = "daraz_reviews.xlsx"
    file_path = os.path.join(current_dir, file_name)
    workbook, worksheet = initialize_workbook(file_path)
    next_row = worksheet.max_row + 1
    product_index = 1
    total_reviews_collected = 0
    signal.signal(signal.SIGINT, lambda sig, frame: signal_handler(sig, frame, driver, workbook, file_path, total_reviews_collected))
    while True:
        driver.get(base_url)
        try:
            wait = WebDriverWait(driver, 10)
            products = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".product-card--vHfY9")))
            if product_index > len(products):
                break
            products[product_index - 1].click()
            time.sleep(2)
            reviews = scroll_to_reviews(driver)
            if not reviews:
                print(f"\rSkipping item {product_index} due to missing reviews.", end="", flush=True)
                product_index += 1
                continue
            total_reviews_collected = collect_reviews(driver, user_input_item, worksheet, max_pages, total_reviews_collected)
            product_index += 1
        except:
            break
    save_reviews_and_exit(driver, workbook, file_path, total_reviews_collected)

if __name__ == "__main__":
    main()